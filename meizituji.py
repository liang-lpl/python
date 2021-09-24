#!/usr/bin/env python
# encoding: utf-8
# author: ricia
import requests
import bs4
import openpyxl
import re

# 发送GET请求
def open_url(url):
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.98 Safari/537.36'}
    res = requests.get(url,headers=headers)
    # 网站编码是‘gb2312’
    res.encoding = 'gb2312'
    return res


def find_names(res):
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    # 文件名
    names = []
    targets1 = soup.find_all('title')
    for each in targets1:
        names.append(each.text)

    # 找出下载地址
    downloads = []
    # 匹配.zip结尾的下载地址
    targets2 = soup.find_all(name='a', attrs={"href": re.compile(r'.zip[/hide]')})
    for download in targets2:
        downloads.append(download.get('href'))

    result = []
    length = len(names)
    for i in range(length):
        result.append([names[i], downloads[i]])

    return result

# 主页找页面数量
def find_depth(res):
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    depth = soup.find('strong').text
    return int(depth)

# 找出一页的超链接
def find_link(res):
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    links = []
    targets2 = soup.find_all(name='a', attrs={"href": re.compile(r'^[url]https://www.ku137.net/b/[/url]\d*/\d*.html')})
    for link in targets2:
        links.append(link.get('href'))

    return links

# 找所有页面超链接
def all_link(host,depth):
    all_links = []
    for i in range(0, depth):
        url = host + 'list_1_' + str(i + 1) + '.html'
        res = open_url(url)
        all_links.extend(find_link(res))

    return all_links

# 保存本地EXCEL文件
def save_to_excel(result):
    wb = openpyxl.Workbook()
    # 获取活跃的工作表
    ws = wb.active
    # 给指定单位格赋值
    ws['A1'] = "文件名"
    ws['B1'] = "下载地址"

    # 逐行添加文件名和下载地址
    for each in result:
        ws.append(each)

    # 保存文件
    wb.save("deom.xlsx")

def main():
    host = 'https://www.ku137.net/b/1/'
    res = open_url(host)
    depth = find_depth(res)
    all_links = all_link(host,depth)
    result = []
    for each_link in all_links:
        result.extend(find_names(open_url(each_link)))

    save_to_excel(result)

if __name__ == "__main__":
    main()